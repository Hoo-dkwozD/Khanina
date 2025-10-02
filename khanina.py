#!/usr/bin/env python3

# Standard library
import argparse
from datetime import datetime
import json
from pathlib import Path
import pprint
import re
import time

# Third-party libraries
import colorama
from colorama import Fore, Style
from jinja2.sandbox import SandboxedEnvironment
from jsonpointer import JsonPointer, JsonPointerException
import litellm
from openpyxl import load_workbook, Workbook
import requests
from tqdm import tqdm

# Initialize colorama
colorama.init()

def print_info(msg):
    print(f"{Fore.BLUE}[+] {msg}{Style.RESET_ALL}")

def print_warning(msg):
    print(f"{Fore.YELLOW}[!] {msg}{Style.RESET_ALL}")

def print_error(msg):
    print(f"{Fore.RED}[!] {msg}{Style.RESET_ALL}")

def print_success(msg):
    print(f"{Fore.GREEN}[+] {msg}{Style.RESET_ALL}")

def ascii_art():
    art = """
  ____  __.__                  .__               
 |    |/ _|  |__ _____    ____ |__| ____ _____   
 |      < |  |  \\__  \\  /    \\|  |/    \\__  \\  
 |    |  \\|   Y  \\/ __ \\|   |  \\  |   |  \\/ __ \\_
 |____|__ \\___|  (____  /___|  /__|___|  (____  /
         \\/    \\/     \\/     \\/        \\/     \\/ 

    """
    print(Fore.GREEN + art + Style.RESET_ALL)

# Main function
def main():
    parser = argparse.ArgumentParser(
        description="Khaniña - LLM Prompt Injection Fuzzer",
        epilog="This tool sends prompts from Excel files to an LLM API endpoint and collects responses for analysis. Use verbose mode for detailed output during prompt processing."
    )
    parser.add_argument('-v', '--verbose', action='store_true', help='Enable verbose output, showing detailed progress for each prompt including index out of total and response times')
    args = parser.parse_args()

    ascii_art()
    print_info("Welcome to Khaniña - LLM Prompt Injection Fuzzer")

    # Check for config files
    headers_file = Path('resources/headers.json')
    body_file = Path('resources/body.json')

    if not headers_file.exists() or not body_file.exists():
        print_warning("Configuration files not found. Entering preparation phase.")
        print_info("Please copy resources/headers.json.example to resources/headers.json and modify with your endpoint details.")
        print_info("Please copy resources/body.json.example to resources/body.json and modify with your request body (or leave empty for GET).")
        print_info("Also, place your prompt Excel files in the 'prompts' directory.")
        return

    # Load configs
    try:
        with open(headers_file) as f:
            headers_config = json.load(f)
        with open(body_file) as f:
            body_config = json.load(f)
    except json.JSONDecodeError as e:
        print_error(f"Invalid JSON in config files: {e}")
        return

    # Validate essential fields
    if 'base_url' not in headers_config:
        print_error("headers.json must contain 'base_url' field.")
        return
    if not headers_config['base_url'].startswith('http') and not headers_config['base_url'].startswith('https'):
        print_error("The 'base_url' in headers.json must be a URL starting with http or https.")
        return
    if 'endpoint' not in headers_config:
        print_error("headers.json must contain 'endpoint' field.")
        return
    if not headers_config['endpoint'].startswith('/'):
        print_error("The 'endpoint' in headers.json must start with '/'.")
        return
    if 'method' not in headers_config:
        print_error("headers.json must contain 'method' field.")
        return
    if headers_config['method'].upper() not in ['GET', 'POST', 'PUT', 'DELETE', 'PATCH']:
        print_error("Invalid HTTP method in headers.json.")
        return
    if not isinstance(headers_config.get('headers', {}), dict):
        print_error("'headers' in headers.json must be a JSON object.")
        return
    if "{{PROMPT}}" not in json.dumps(body_config):
        print_warning("body.json does not contain '{{PROMPT}}' placeholder. Ensure your prompt is inserted correctly.")
        # Choice to continue or not
        cont = input("Do you want to continue? (y/n): ").strip().lower()
        if cont != 'y':
            print_info("Operation cancelled.")
            return

    if args.verbose:
        print_info("Configuration files loaded successfully.")

    # Display details
    print_info("Endpoint Details:")
    print(f"  Method: {headers_config.get('method')}")
    print(f"  Base URL: {headers_config.get('base_url')}")
    print(f"  Endpoint: {headers_config.get('endpoint')}")
    print(f"  Headers: {json.dumps(headers_config.get('headers', {}), indent=2)}")
    print_info("Request Body:")
    print(f"  {json.dumps(body_config, indent=2)}")

    confirm = input("Do you want to proceed with these settings? (y/n): ").strip().lower()
    if confirm != 'y':
        print_info("Operation cancelled.")
        return

    # Send test request
    print_info("Sending test request to validate configuration...")
    test_url = headers_config.get('base_url') + headers_config.get('endpoint')
    test_headers = headers_config.get('headers')
    test_body = body_config.copy()
    # Insert a test prompt
    test_body_str = json.dumps(test_body)
    test_body_str = test_body_str.replace("{{PROMPT}}", "test")
    try:
        test_body = json.loads(test_body_str)
    except json.JSONDecodeError:
        print_error("Error inserting test prompt into request body.")
        return
    try:
        method = headers_config.get('method').upper()
        if method == 'GET':
            test_response = requests.get(test_url, headers=test_headers, params=test_body)
        else:
            test_response = requests.request(method, test_url, headers=test_headers, data=json.dumps(test_body, separators=(',', ':')))
        if test_response.status_code == 200:
            print_success("Test request successful. Endpoint is reachable.")
        else:
            print_warning(f"Test request returned status code {test_response.status_code}. Check your configuration.")
            # Print out request and response details for debugging
            print_info(f"Request URL: {test_response.request.url}")
            print_info(f"Request Headers: {test_response.request.headers}")
            print_info(f"Request Body: {test_response.request.body}")
            print_info(f"Response Body: {test_response.text}")

            proceed = input("Do you want to proceed anyway? (y/n): ").strip().lower()
            if proceed != 'y':
                print_info("Operation cancelled.")
                return
    except Exception as e:
        print_error(f"Error sending test request: {e}")
        return
    
    if args.verbose:
        print_info("Test request completed.")
    print_info("Entering main operation phase.")

    # LLM Evaluation Option
    enable_llm_eval = input("Do you want to enable LLM evaluation for jailbreak detection? (y/n): ").strip().lower() == 'y'
    llm_config = None
    eval_template = None
    llm_purpose = None
    if enable_llm_eval:
        # Check for llm.json
        llm_config_file = Path('resources/llm.json')
        if not llm_config_file.exists():
            print_warning("llm.json not found. Please copy resources/llm.json.example to resources/llm.json and configure.")
            return

        # Load llm config
        try:
            with open(llm_config_file) as f:
                llm_config = json.load(f)
        except json.JSONDecodeError as e:
            print_error(f"Invalid JSON in llm.json: {e}")
            return

        # Validate provider
        supported_providers = ['openai', 'cohere', 'anthropic', 'claude', 'gemini', 'deepseek', 'ollama']
        provider = llm_config.get('provider', '').lower()
        if provider not in supported_providers:
            print_error(f"Unsupported provider: {provider}. Supported: {', '.join(supported_providers)}")
            return

        # Check required keys
        if 'model' not in llm_config:
            print_error("llm.json must contain 'model'.")
            return
        if provider != 'ollama' and 'api_key' not in llm_config:
            print_error("llm.json must contain 'api_key' for this provider.")
            return

        # Load evaluate.prompt
        eval_prompt_file = Path('resources/evaluate.prompt')
        if not eval_prompt_file.exists():
            print_error("evaluate.prompt not found in resources/.")
            return
        with open(eval_prompt_file) as f:
            eval_template_str = f.read()

        # Load llm.purpose
        llm_purpose_file = Path('resources/llm.purpose')
        if not llm_purpose_file.exists():
            print_error("llm.purpose not found in resources/. Please copy resources/llm.purpose.example to resources/llm.purpose and configure.")
            return
        with open(llm_purpose_file) as f:
            llm_purpose = f.read().strip()
        if llm_purpose == "":
            print_warning("llm.purpose is empty. Add a purpose for proper evaluation context.")
            return

        # Create Jinja2 environment
        jinja_env = SandboxedEnvironment()
        eval_template = jinja_env.from_string(eval_template_str)

        print_info("LLM evaluation enabled.")

    # Project management
    results_dir = Path('results')
    results_dir.mkdir(exist_ok=True)
    existing_projects = [d.name for d in results_dir.iterdir() if d.is_dir()]

    if existing_projects:
        print_info("Existing projects:")
        for i, proj in enumerate(existing_projects, 1):
            print(f"  {i}. {proj}")
        choice = input("Choose an existing project (number) or 'n' for new: ").strip()
        if choice.lower() == 'n':
            project_name = input("Enter new project name (lowercase alphanum only): ").strip()
            if not re.match(r'^[a-z0-9]+$', project_name):
                print_error("Invalid project name. Must be lowercase alphanum.")
                return
        else:
            try:
                project_name = existing_projects[int(choice) - 1]
            except (ValueError, IndexError):
                print_error("Invalid choice.")
                return
    else:
        project_name = input("Enter new project name (lowercase alphanum only): ").strip()
        if not re.match(r'^[a-z0-9]+$', project_name):
            print_error("Invalid project name. Must be lowercase alphanum.")
            return

    project_dir = results_dir / project_name
    project_dir.mkdir(exist_ok=True)

    # JSON pointer
    pointer = input("Enter JSON pointer for main key-value extraction (e.g., /choices/0/message/content): ").strip()

    # Prompt selection
    prompts_dir = Path('prompts')
    if not prompts_dir.exists():
        print_error("Prompts directory not found.")
        return

    prompt_files = list(prompts_dir.glob('*.xlsx'))
    if not prompt_files:
        print_error("No Excel files found in prompts directory.")
        return

    print_info("Available prompt files:")
    for i, pf in enumerate(prompt_files, 1):
        print(f"  {i}. {pf.name}")
        # Display first few prompts
        try:
            wb = load_workbook(pf)
            ws = wb.active
            if ws is not None:
                prompts = [ws.cell(row=r, column=2).value for r in range(2, min(6, ws.max_row + 1)) if ws.cell(row=r, column=2).value]
                print(f"    Sample prompts: {', '.join(str(p)[:50] for p in prompts[:3])}")
            else: 
                print_warning(f"No active sheet in {pf.name}")
        except Exception as e:
            print_warning(f"Error reading {pf.name}: {e}")

    selected_indices = input("Enter file numbers to use (comma-separated, or 'all'): ").strip()
    if selected_indices.lower() == 'all':
        selected_files = prompt_files
    else:
        try:
            indices = [int(x.strip()) - 1 for x in selected_indices.split(',')]
            selected_files = [prompt_files[i] for i in indices if 0 <= i < len(prompt_files)]
        except ValueError:
            print_error("Invalid selection.")
            return

    if enable_llm_eval:
        success_col = []

    # Process each selected file
    for pf in selected_files:
        print_info(f"Processing {pf.name}")
        try:
            wb = load_workbook(pf)
            ws = wb.active
            if ws is not None:
                prompts = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=2).value]
            else:
                print_warning(f"No active sheet in {pf.name}, skipping.")
                continue
        except Exception as e:
            print_error(f"Error reading prompts from {pf.name}: {e}")
            continue

        # Create results Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        result_filename = f"{pf.stem} - {project_name} - {timestamp} - Results.xlsx"
        result_path = project_dir / result_filename
        result_wb = Workbook()
        result_ws = result_wb.active
        if result_ws is not None:
            if enable_llm_eval:
                result_ws.append(['Index', 'Prompt', 'Main', 'Full Response', 'Response Time (s)', 'Success', 'Confidence', 'Evaluator Response'])
            else:
                result_ws.append(['Index', 'Prompt', 'Main', 'Full Response', 'Response Time (s)'])
        else:
            print_error(f"Failed to create results worksheet for {pf.name}, skipping.")
            continue

        if not args.verbose:
            pbar = tqdm(total=len(prompts), desc="Processing prompts")

        for idx, prompt in enumerate(prompts, 1):
            # Ensure prompt is string
            if not isinstance(prompt, str):
                prompt = str(prompt)

            # Strip whitespace and quotation marks
            prompt = prompt.strip().strip('"').strip("'")

            # Skip empty prompts
            if prompt == "":
                if args.verbose:
                    print_warning(f"Skipping empty prompt at index {idx}.")
                continue

            if args.verbose:
                print_info(f"Sending prompt {idx}/{len(prompts)}: {prompt[:50]}...")

            # Prepare request
            method = headers_config.get('method').upper()
            url = headers_config.get('base_url') + headers_config.get('endpoint')
            headers = headers_config.get('headers', {})
            body = body_config.copy()
            # Place prompt in body if placeholder exists, replace all occurrences, at least once
            body_str = json.dumps(body)
            body_str = body_str.replace("{{PROMPT}}", prompt)
            try:
                body = json.loads(body_str)
            except json.JSONDecodeError:
                print_error("Error inserting prompt into request body.")
                continue

            start_time = time.time()
            try:
                if method == 'GET':
                    response = requests.get(url, headers=headers, params=body)
                else:
                    response = requests.request(method, url, headers=headers, data=json.dumps(body, separators=(',', ':')))
                response_time = time.time() - start_time

                if response.status_code == 200:
                    try:
                        resp_json = response.json()
                        main_value = JsonPointer(pointer).get(resp_json)
                        full_resp = json.dumps(resp_json)

                        if args.verbose:
                            print_info(f"Extracted main value: {main_value}")
                    except (JsonPointerException, json.JSONDecodeError):
                        main_value = "Error extracting"
                        full_resp = response.text
                else:
                    main_value = f"HTTP {response.status_code}"
                    full_resp = response.text

                # LLM Evaluation
                if enable_llm_eval and eval_template is not None and llm_config is not None:
                    try:
                        print_info(f"Evaluating response with {provider} LLM...")
                        eval_prompt_rendered = eval_template.render(original_prompt=prompt, response=main_value, llm_purpose=llm_purpose)
                        response_eval = None
                        response_eval = litellm.completion(
                            model=llm_config['model'],
                            messages=[{"role": "user", "content": eval_prompt_rendered}],
                            api_key=llm_config.get('api_key'),
                        )
                        eval_content = response_eval.choices[0].message.content.strip() # type: ignore
                        eval_content = eval_content.strip('`') # Remove backticks if any
                        eval_data = json.loads(eval_content)
                        success = eval_data.get('success', False)
                        confidence = eval_data.get('confidence', 0)
                        evaluator_response = eval_content
                    except Exception as e:
                        print_warning(f"Evaluation failed for prompt {idx}: {e}")
                        success = "Error"
                        confidence = "Error"
                        evaluator_response = str(e)
                        if response_eval is not None:
                            evaluator_response += " | LLM response: " + str(response_eval)
                        if args.verbose:
                            print_warning(f"Evaluation response: {evaluator_response}")
                else:
                    success = None
                    confidence = None
                    evaluator_response = None

                if enable_llm_eval:
                    result_ws.append([idx, prompt, str(main_value), full_resp, response_time, success, confidence, evaluator_response])

                    if success:
                        success_col.append(
                            {'Index': idx, 'Prompt': prompt, 'Main': main_value, 'Response Time (s)': response_time, 'Success': success, 'Confidence': confidence, 'Evaluator Response': evaluator_response}
                        )
                else:
                    result_ws.append([idx, prompt, str(main_value), full_resp, response_time])

                if args.verbose:
                    print_success(f"Response received in {response_time:.2f}s")

            except Exception as e:
                print_error(f"Error sending request for prompt {idx}: {e}")
                if enable_llm_eval:
                    result_ws.append([idx, prompt, "Error", str(e), 0, "N/A", "N/A", "N/A"])
                else:
                    result_ws.append([idx, prompt, "Error", str(e), 0])

            if not args.verbose:
                pbar.update(1)

        if not args.verbose:
            pbar.close()

        result_wb.save(result_path)
        print_success(f"Results saved to {result_path}")

    if enable_llm_eval and len(success_col) > 0:
        # Display summary of successful jailbreaks
        print_info("Summary of successful jailbreaks:")
        for entry in success_col:
            print(f"  Index: {entry['Index']}, Prompt: {entry['Prompt'][:50]}..., Main: {entry['Main']}, Confidence: {entry['Confidence']}")

    print_success("All processing complete. Goodbye!")

if __name__ == "__main__":
    main()
